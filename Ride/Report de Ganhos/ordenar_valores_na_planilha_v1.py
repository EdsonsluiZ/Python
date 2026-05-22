# -*- coding: utf-8 -*-

__author__ = 'Edson Luiz'

from openpyxl import load_workbook
from datetime import datetime
import re

arquivo = r"C:/Python/Ride/Report de Ganhos/Relatório Ganhos Uber Driver.xlsx"
saida = r"C:/Python/Ride/Report de Ganhos/Relatorio_Ganhos_Uber_Driver_Sheet4.xlsx"

wb = load_workbook(arquivo)
ws3 = wb.worksheets[2]

# Remove Sheet4 se já existir
if "Sheet4" in wb.sheetnames:
    del wb["Sheet4"]

ws4 = wb.create_sheet("Sheet4")

ws4.append(["Data", "Tipo", "Hora", "Valor"])

ano_padrao = 2026

meses = {
    "jan": 1, "jan.": 1, "janeiro": 1,
    "fev": 2, "fev.": 2, "fevereiro": 2,
    "mar": 3, "mar.": 3, "março": 3, "marco": 3,
    "abr": 4, "abr.": 4, "abril": 4,
    "mai": 5, "mai.": 5, "maio": 5,
    "jun": 6, "jun.": 6, "junho": 6,
    "jul": 7, "jul.": 7, "julho": 7,
    "ago": 8, "ago.": 8, "agosto": 8,
    "set": 9, "set.": 9, "setembro": 9,
    "out": 10, "out.": 10, "outubro": 10,
    "nov": 11, "nov.": 11, "novembro": 11,
    "dez": 12, "dez.": 12, "dezembro": 12,
}

padrao_data = re.compile(r"^(\d{1,2})\s+de\s+([a-zçã.]+)$", re.IGNORECASE)
padrao_hora = re.compile(r"^\d{1,2}:\d{2}:\d{2}$")
padrao_valor = re.compile(r"^-?\d+(?:\.\d+)?$")

data_atual = None
tipo_atual = None
hora_atual = None

registros = []

for linha in range(1, ws3.max_row + 1):

    valor = ws3.cell(row=linha, column=1).value

    if valor is None:
        continue

    texto = str(valor).strip()

    if not texto:
        continue

    # Detecta Data
    m_data = padrao_data.match(texto.lower())

    if m_data:

        dia = int(m_data.group(1))
        mes_txt = m_data.group(2).lower()

        if mes_txt in meses:

            data_atual = datetime(
                ano_padrao,
                meses[mes_txt],
                dia
            )

            tipo_atual = None
            hora_atual = None

        continue

    # Detecta Tipo
    if (
        data_atual is not None
        and not padrao_hora.match(texto)
        and not padrao_valor.match(texto)
    ):

        tipo_atual = texto
        hora_atual = None
        continue

    # Detecta Hora
    if (
        data_atual is not None
        and tipo_atual is not None
        and padrao_hora.match(texto)
    ):

        hora_atual = texto
        continue

    # Detecta Valor
    if (
        data_atual is not None
        and tipo_atual is not None
        and hora_atual is not None
        and padrao_valor.match(texto)
    ):

        valor_float = float(texto)

        registros.append([
            data_atual,
            tipo_atual,
            hora_atual,
            valor_float
        ])

        tipo_atual = None
        hora_atual = None


# =========================================================
# ORDENAÇÃO DA DATA MAIS ANTIGA PARA A MAIS RECENTE
# =========================================================

from datetime import datetime

registros.sort(
    key=lambda x: (
        x[0],  # data
        datetime.strptime(x[2], "%H:%M:%S").time()  # hora
    )
)

# =========================================================
# GRAVAÇÃO NA SHEET4
# =========================================================

for registro in registros:
    ws4.append(registro)


# =========================================================
# FORMATAÇÃO
# =========================================================

for cell in ws4["A"]:
    cell.number_format = "dd/mm/yyyy"

for cell in ws4["D"]:
    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

ws4.column_dimensions["A"].width = 15
ws4.column_dimensions["B"].width = 35
ws4.column_dimensions["C"].width = 12
ws4.column_dimensions["D"].width = 18


# =========================================================
# SALVAR
# =========================================================

wb.save(saida)

print("Processo concluído com sucesso.")
print(f"Registros transportados: {len(registros)}")
print(f"Arquivo salvo em: {saida}")