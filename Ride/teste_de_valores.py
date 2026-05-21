# -*- coding: utf-8 -*-

__author__ = 'Edson Luiz'

from openpyxl import load_workbook

arquivo = r"C:/Python/Ride/Report de Ganhos/Relatório Ganhos Uber Driver.xlsx"

wb = load_workbook(arquivo)
ws3 = wb.worksheets[2]

for linha in range(1, 80):
    valores = []
    for col in range(1, ws3.max_column + 1):
        valor = ws3.cell(row=linha, column=col).value
        if valor is not None and str(valor).strip() != "":
            valores.append(str(valor).strip())

    if valores:
        print(linha, valores)